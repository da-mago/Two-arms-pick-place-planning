###############################################################################
#
# Model containing required YuMi stuff for solving pick&place problem:
# - LUTs (excel format):
#   - collisions (including minimun distance between arms)
#   - arms position reachability
#   - grid arms configuration (angles)
#
# - Methods:
#   - isThereCollision()
#   - isReachable()
#   - plot()
#
#   CERRADO: solo valido para Yumi
#
###############################################################################

import numpy as np
from openpyxl import load_workbook

class YuMi():
    ''' YuMi (14 degrees of freedom) robot. '''

    def __init__(self):
        self.safe_distance = 1
        self.M, self.N, self.collision, self.config, self.reachable, self.location = self._load_spreadsheet()    

    def _load_spreadsheet(self):

        wb = load_workbook(filename = 'Yumi_IRB1400.xlsx')
        ws0 = wb['configuration']
        ws1 = wb['collision']

        M  = ws0.cell(row=1, column=2).value # Grid M*N
        N  = ws0.cell(row=1, column=3).value
        Zs = ws0.cell(row=1, column=5).value # Number of Z planes

        NUM_ARMS   = 2
        NUM_ANGLES = 7
        NUM_DIM    = 3 # XYZ
        location  = np.zeros((Zs, N, M, NUM_DIM), dtype=np.int16)
        reachable = np.zeros((Zs, NUM_ARMS, N, M), dtype=np.uint8)
        config    = np.zeros((Zs, NUM_ARMS, N, M, NUM_ANGLES), dtype=float)

        for i in range(Zs):
            base = i*(M*N + 2)
            row_init = base + 3
            z = int((ws0.cell(row=row_init, column=COL_R_REACH).value).split('=')[1]) # "Alcanza Z=180" -> 180
            row_init = base + 4
            location[i]     = np.array([[row[0].value, row[1].value, z] for row in ws0.iter_rows(min_row=row_init, min_col=1, max_row=(row_init+M*N-1), max_col=2)]).reshape((N, M, NUM_DIM))
            for arm in range(NUM_ARMS):
                col_r_reach = 3 + (NUM_ANGLES + 3)*i
                col_r_ang   = col_r_reach + 1
                reachable[i][arm] = np.array([ row[0].value for row in ws0.iter_rows(min_row=row_init, min_col=col_r_reach, max_row=(row_init+M*N-1), max_col=col_r_reach)]).reshape((N, M))
                config[i][arm]    = np.array([[col.value for col in row] for row in ws0.iter_rows(min_row=row_init, min_col=col_r_ang, max_row=(row_init+M*N-1), max_col=(col_r_ang+NUM_ANGLES-1))]).reshape((N, M, NUM_ANGLES))

        COL_COLLISION = 5
        MIN_DISTANCE = 29
        reshape = (N,M)*NUM_ARMS
        collision = np.array([[(row[0].value * row[1].value) >= MIN_DISTANCE] for row in ws1.iter_rows(min_row=2, min_col=COL_COLLISION, max_row=2+((M*N)**NUM_ARMS)-1, max_col=(COL_COLLISION+1))], dtype=np.uint8).reshape(reshape)

        wb.close()

        return M, N, collision, config, reachable, location


    def checkCollision(self, armsGridPos):
        ''' Check Robot collision.

            Note: any distance lower than 'collision_distance' is also 
                  considered collision.
        '''
        # Check also EEs distance
        (x1,y1),(x2,y2) = armsGridPos
        
        return self.collision[y1, x1, y2, x2]


    def checkReachability(self, armsGridpos):
        ''' Check that both arms can reach their target positions '''
        for i,(x,y) in enumerate(armsGridPos):
            if self.reachable[0, i, y, x]:
                return False

        return True

    def validate(self, armsGridPos):
        # check:
        # - collision
        # - reachability
        # - Grid boundaries
        return True

    def plot(self, armsGridPos, ax, plt3d):
        ''' Draw only EEs position (just a circle) '''
        for col,(x,y) in zip(['r', 'b'], armsGridPos):
            if plt3d: ax.plot3D(x, y, z, c=col, marker='o')
            else:     ax.plot  (x, y,    c=col, marker='o')


if __name__ == "__main__":

    # Robot
    robot = YuMi()
